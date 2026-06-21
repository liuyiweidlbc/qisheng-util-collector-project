# -*- coding: utf-8 -*-
"""生成初一寒假作业计划统筹 Excel 模板"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 输出文件路径（英文文件名避免编码问题，可自行重命名为「初一寒假作业计划统筹.xlsx」）
import os
_here = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(_here, "winter_homework_plan.xlsx")

# 学科列表
SUBJECTS = ["数学", "英语", "语文", "生物", "历史", "地理", "道法"]

# 学科表头
SUBJECT_HEADERS = ["作业项", "计划开始", "计划时长", "实际开始", "实际时长", "是否完成(✔)"]


def style_header(ws, row=1):
    """表头样式"""
    thin = Side(style="thin", color="000000")
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")


def style_cell_border(ws, range_str=None):
    """给区域加边框"""
    thin = Side(style="thin", color="000000")
    for row in ws.iter_rows(range_str or ws.dimensions):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_daily_schedule_sheet(wb):
    """第一个 sheet：每日时间。A列 2/2-2/20 日期，第1行从 8:00 递增到 22:00"""
    ws = wb.create_sheet("每日时间", 0)
    # 日期：2月2日 至 2月20日（共19天）
    base_year = 2025
    for i in range(19):
        d = datetime(base_year, 2, 2 + i)
        ws.cell(row=i + 2, column=1, value=d.strftime("%m/%d"))
    ws.cell(row=1, column=1, value="日期")
    # 第1行第2列起：8:00 到 22:00，每小时一列
    for h in range(8, 23):
        ws.cell(row=1, column=h - 8 + 2, value=f"{h:02d}:00")
    # 表头样式
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 17):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")
    # 列宽
    ws.column_dimensions["A"].width = 8
    for c in range(2, 17):
        ws.column_dimensions[get_column_letter(c)].width = 6
    thin = Side(style="thin", color="000000")
    for r in range(1, 21):
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if r > 1 and c > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def build_subject_sheet(wb, name):
    """学科 sheet：作业项、计划开始、计划时长、实际开始、实际时长、是否完成(✔)"""
    ws = wb.create_sheet(name)
    for col, h in enumerate(SUBJECT_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1)
    # 列宽
    ws.column_dimensions["A"].width = 28
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12
    # 预留几行空行便于填写
    thin = Side(style="thin", color="000000")
    for r in range(2, 22):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def build_summary_sheet(wb):
    """汇总 sheet：已完成所用时间、未完成所需时间"""
    ws = wb.create_sheet("汇总")
    ws.cell(row=1, column=1, value="统计项")
    ws.cell(row=1, column=2, value="时间")
    ws.cell(row=2, column=1, value="已完成所用时间")
    ws.cell(row=3, column=1, value="未完成所需时间")
    ws.cell(row=2, column=2, value="")  # 可填或公式
    ws.cell(row=3, column=2, value="")
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    thin = Side(style="thin", color="000000")
    for r in range(1, 4):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")


def main():
    wb = Workbook()
    # 删除默认的 Sheet（若有），由我们按顺序创建
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_daily_schedule_sheet(wb)
    for subj in SUBJECTS:
        build_subject_sheet(wb, subj)
    build_summary_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
